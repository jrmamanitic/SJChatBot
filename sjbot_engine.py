# ==============================================================
# SJBot — Motor (backend)
# Steve Jobs College - Tacna | Chatbot institucional y académico
# ==============================================================
import io
import os
import re
import subprocess
import tempfile
import unicodedata
import warnings
from pathlib import Path

import openpyxl
import streamlit as st

warnings.filterwarnings("ignore")

BASE_DIR = Path(__file__).parent
RUTA_PLANTILLA_LOCAL = BASE_DIR / "assets" / "PLANTILLA_SECUNDARIA.xlsx"

_PAT_IDX = re.compile(r"CONSOLIDADO!\$([A-Z]{1,2})\$1(?!:)")
_PAT_GRADO = re.compile(r"([1-5])\s*(?:RO|DO|ER|TO|ERO|°|º)?\s*A[ÑN]O", re.I)
_PAT_BIM = re.compile(r"\b(I{1,3}|IV)\s*BIM", re.I)


def _norm(s):
    s = unicodedata.normalize("NFD", str(s))
    return "".join(c for c in s if unicodedata.category(c) != "Mn").lower().strip()


def _nfc(s):
    return unicodedata.normalize("NFC", str(s))


# ==============================================================
# 1) ACCESO A ARCHIVOS: Google Drive (service account) o subida manual
# ==============================================================

@st.cache_resource(show_spinner=False)
def _drive_service():
    """Crea el cliente de Google Drive a partir de las credenciales en st.secrets."""
    from google.oauth2 import service_account
    from googleapiclient.discovery import build

    info = dict(st.secrets["gcp_service_account"])
    creds = service_account.Credentials.from_service_account_info(
        info, scopes=["https://www.googleapis.com/auth/drive.readonly"]
    )
    return build("drive", "v3", credentials=creds, cache_discovery=False)


@st.cache_data(ttl=300, show_spinner="Leyendo registros desde Google Drive…")
def cargar_archivos_drive(folder_id: str) -> dict:
    """Recorre recursivamente la carpeta de Drive y devuelve {nombre_archivo: bytes}."""
    service = _drive_service()
    archivos = {}

    def _walk(fid):
        page_token = None
        while True:
            resp = service.files().list(
                q=f"'{fid}' in parents and trashed=false",
                fields="nextPageToken, files(id, name, mimeType)",
                pageToken=page_token,
            ).execute()
            for f in resp.get("files", []):
                if f["mimeType"] == "application/vnd.google-apps.folder":
                    _walk(f["id"])
                elif f["name"].lower().endswith(".xlsx") and not f["name"].startswith("~$"):
                    data = service.files().get_media(fileId=f["id"]).execute()
                    archivos[f["name"]] = data
            page_token = resp.get("nextPageToken")
            if not page_token:
                break

    _walk(folder_id)
    return archivos


def cargar_archivos_subidos(files) -> dict:
    """Convierte los archivos subidos manualmente (st.file_uploader) al mismo formato {nombre: bytes}."""
    return {f.name: f.getvalue() for f in files}


# ==============================================================
# 2) ÍNDICE DE REGISTROS (a partir del dict {nombre: bytes})
# ==============================================================

def _wb(nombre_o_bytes, archivos=None, data_only=True, read_only=False):
    """Abre un workbook ya sea desde disco (plantilla local) o desde bytes en memoria."""
    if archivos is not None:
        return openpyxl.load_workbook(io.BytesIO(archivos[nombre_o_bytes]), data_only=data_only, read_only=read_only)
    return openpyxl.load_workbook(nombre_o_bytes, data_only=data_only, read_only=read_only)


def indexar_registros(archivos: dict) -> dict:
    """{grado: {bimestre: nombre_archivo}} a partir de los .xlsx disponibles."""
    regs, errores, sin_match = {}, [], []
    for nombre in archivos:
        if "PLANTILLA" in nombre.upper():
            continue
        nfc = _nfc(nombre)
        mg = _PAT_GRADO.search(nfc)
        if not mg:
            sin_match.append(nombre)
            continue
        try:
            wb = _wb(nombre, archivos, read_only=True)
            if "CONSOLIDADO" not in wb.sheetnames:
                continue
        except Exception as e:
            errores.append(f"{nombre}: {e}")
            continue
        mb = _PAT_BIM.search(nfc)
        grado = f"{mg.group(1)}° AÑO"
        bim = (mb.group(1).upper() + " BIM") if mb else "BIM"
        regs.setdefault(grado, {})[bim] = nombre
    regs = dict(sorted(regs.items()))
    return {"registros": regs, "errores": errores, "sin_match": sin_match}


def _leer_alumnos(nombre_reg, archivos):
    cons = _wb(nombre_reg, archivos)["CONSOLIDADO"]
    alumnos = {}
    for r in range(5, cons.max_row + 1):
        n, nom = cons.cell(r, 1).value, cons.cell(r, 2).value
        if n and nom and str(nom) != "0":
            try:
                alumnos[int(float(n))] = str(nom).strip()
            except (ValueError, TypeError):
                pass
    return alumnos


def catalogo(archivos):
    idx = indexar_registros(archivos)["registros"]
    out = {}
    for grado, bims in idx.items():
        out[grado] = {b: _leer_alumnos(r, archivos) for b, r in bims.items()}
    return out


def _match_grado(texto, archivos):
    regs = indexar_registros(archivos)["registros"]
    t = _norm(_nfc(texto))
    ordinales = {"primer": "1", "segund": "2", "tercer": "3", "cuart": "4", "quint": "5"}
    m = re.search(r"[1-5]", t)
    num = m.group() if m else next((v for k, v in ordinales.items() if k in t), None)
    if num:
        for g in regs:
            if g.startswith(num):
                return g
    return None


def buscar_alumno(consulta, archivos, grado=None, bimestre=None):
    regs = indexar_registros(archivos)["registros"]
    grados = [_match_grado(grado, archivos)] if grado and _match_grado(grado, archivos) else list(regs)
    q = _norm(consulta)
    res = []
    for g in grados:
        if g not in regs:
            continue
        bims = regs[g]
        bsel = None
        if bimestre:
            for b in bims:
                if _norm(bimestre).replace(" ", "") in _norm(b).replace(" ", ""):
                    bsel = b
        bsel = bsel or sorted(bims)[-1]
        alumnos = _leer_alumnos(bims[bsel], archivos)
        if q.isdigit() and int(q) in alumnos:
            res.append((g, bsel, int(q), alumnos[int(q)]))
            continue
        toks = q.split()
        for n, nom in alumnos.items():
            if toks and all(t in _norm(nom) for t in toks):
                res.append((g, bsel, n, nom))
    return res


# ==============================================================
# 3) LIBRETA EN PDF
# ==============================================================

def generar_libreta_pdf(grado, n_lista, archivos, bimestre=None, out_dir=None):
    regs = indexar_registros(archivos)["registros"]
    g = _match_grado(grado, archivos) or grado
    if g not in regs:
        raise ValueError(f"Grado no encontrado: {grado}. Disponibles: {list(regs)}")
    bims = regs[g]
    bim = None
    if bimestre:
        for b in bims:
            if _norm(bimestre).replace(" ", "") in _norm(b).replace(" ", ""):
                bim = b
    bim = bim or sorted(bims)[-1]
    nombre_reg = bims[bim]

    reg_cons = _wb(nombre_reg, archivos)["CONSOLIDADO"]
    fila = n_lista + 4
    nombre = reg_cons.cell(fila, 2).value
    if not nombre or str(nombre) == "0":
        raise ValueError(f"No existe el alumno N° {n_lista} en {g} ({bim})")
    nombre = str(nombre).strip()

    if not RUTA_PLANTILLA_LOCAL.exists():
        raise FileNotFoundError("No se encontró la plantilla PLANTILLA_SECUNDARIA.xlsx empaquetada con la app.")
    wb_out = openpyxl.load_workbook(RUTA_PLANTILLA_LOCAL, data_only=False)
    lib = wb_out["Libreta"]
    lib["C6"] = n_lista
    lib["D5"] = g
    lib["C8"] = _leer_tutor(nombre_reg, archivos) or "-"
    lib["E7"] = bim.replace("BIM", "BIMESTRE") if "BIM" in bim else bim

    for row in lib.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                refs = _PAT_IDX.findall(cell.value)
                if not refs:
                    continue
                idx = reg_cons[refs[-1] + "1"].value
                val = reg_cons.cell(fila, int(idx)).value if idx else None
                cell.value = val if val not in (None, "", "0") else "-"

    if "CONSOLIDADO" in wb_out.sheetnames:
        del wb_out["CONSOLIDADO"]

    # Ajuste a 1 página de ancho SIEMPRE, aunque falte la fuente Calibri/Carlito en el servidor.
    lib.sheet_properties.pageSetUpPr.fitToPage = True
    lib.page_setup.fitToWidth = 1
    lib.page_setup.fitToHeight = 0
    lib.page_setup.scale = None

    out_dir = Path(out_dir or tempfile.mkdtemp())
    gsafe = re.sub(r"\W+", "", g)
    bsafe = re.sub(r"\W+", "", bim)
    nsafe = re.sub(r"\W+", "_", nombre)[:40]
    tmp_xlsx = out_dir / f"libreta_{gsafe}_{bsafe}_{n_lista:02d}_{nsafe}.xlsx"
    wb_out.save(tmp_xlsx)

    subprocess.run(
        ["soffice", "--headless", "--convert-to", "pdf", "--outdir", str(out_dir), str(tmp_xlsx)],
        check=True, capture_output=True, timeout=120,
    )
    pdf_path = tmp_xlsx.with_suffix(".pdf")
    if not pdf_path.exists():
        raise RuntimeError("LibreOffice no generó el PDF (revisa que 'packages.txt' incluya libreoffice-calc).")
    return str(pdf_path), nombre, g, bim


def _leer_tutor(nombre_reg, archivos):
    try:
        wb = _wb(nombre_reg, archivos, read_only=True)
        if "ASISTEN" not in wb.sheetnames:
            return None
        ws = wb["ASISTEN"]
        for r in range(1, 15):
            etiqueta = ws.cell(r, 1).value
            if etiqueta and "TUTOR" in str(etiqueta).upper():
                val = ws.cell(r, 4).value
                return str(val).strip() if val else None
    except Exception:
        pass
    return None


# ==============================================================
# 4) NOTAS POR CURSO, RESUMEN GENERAL Y ASISTENCIA
# ==============================================================

AREAS_HOJA = {
    "Matemática": "MAT", "Ciencias Sociales": "CSOC", "Educación Física": "EFIS",
    "Desarrollo Personal, Ciudadania y Cívica": "DPCC", "Arte y Cultura": "ARTE",
    "Comunicación": "COMU", "Ciencia y Tecnología": "CTEC",
    "Inglés como lengua Extranjera": "ING", "Educación Religiosa": "REL",
    "Educación por el Trabajo": "EMPR", "Taller de Cómputo": "COMPU",
    "Taller Programación y Robótica": "ROBOTI",
}

_SINONIMOS_CURSO = {
    "matematica": "Matemática", "mate": "Matemática",
    "sociales": "Ciencias Sociales", "ccss": "Ciencias Sociales", "historia": "Ciencias Sociales",
    "educacion fisica": "Educación Física", "efis": "Educación Física", "deporte": "Educación Física",
    "dpcc": "Desarrollo Personal, Ciudadania y Cívica", "desarrollo personal": "Desarrollo Personal, Ciudadania y Cívica",
    "arte": "Arte y Cultura", "cultura": "Arte y Cultura",
    "comunicacion": "Comunicación", "lenguaje": "Comunicación",
    "ciencia": "Ciencia y Tecnología", "ciencia y tecnologia": "Ciencia y Tecnología", "ctec": "Ciencia y Tecnología",
    "ingles": "Inglés como lengua Extranjera", "english": "Inglés como lengua Extranjera",
    "religion": "Educación Religiosa",
    "trabajo": "Educación por el Trabajo", "emprendimiento": "Educación por el Trabajo",
    "computo": "Taller de Cómputo", "computacion": "Taller de Cómputo", "informatica": "Taller de Cómputo",
    "robotica": "Taller Programación y Robótica", "programacion": "Taller Programación y Robótica",
}


def _match_area(texto):
    t = _norm(texto)
    if t in _SINONIMOS_CURSO:
        return _SINONIMOS_CURSO[t]
    for area in AREAS_HOJA:
        if t in _norm(area) or _norm(area) in t:
            return area
    for k, v in _SINONIMOS_CURSO.items():
        if k in t or t in k:
            return v
    return None


_MAPA_AREAS = None


def _construir_mapa_areas():
    global _MAPA_AREAS
    if _MAPA_AREAS is not None:
        return _MAPA_AREAS
    lib = openpyxl.load_workbook(RUTA_PLANTILLA_LOCAL, data_only=False)["Libreta"]
    cons_idx = openpyxl.load_workbook(RUTA_PLANTILLA_LOCAL, data_only=True)["CONSOLIDADO"]
    mapa = {}
    area_actual = None
    for r in range(9, 82):
        b = lib.cell(r, 2).value
        c = lib.cell(r, 3).value
        if isinstance(b, str) and b in AREAS_HOJA:
            area_actual = b
        if area_actual not in AREAS_HOJA:
            continue
        cols = []
        for col_letra in ("D", "E", "F"):
            f = lib[f"{col_letra}{r}"].value
            if isinstance(f, str) and f.startswith("="):
                refs = _PAT_IDX.findall(f)
                idx = cons_idx[refs[-1] + "1"].value if refs else None
                cols.append(int(idx) if idx else None)
            else:
                cols.append(None)
        if c and any(cols):
            mapa.setdefault(area_actual, []).append((str(c).strip(), *cols))
    _MAPA_AREAS = mapa
    return mapa


_ESCALA = {"C": 1, "B": 2, "A": 3, "AD": 4}
_ESCALA_INV = {1: "C (en inicio)", 2: "B (en proceso)", 3: "A (logrado)", 4: "AD (logro destacado)"}


def notas_por_curso(estudiante, curso, archivos, grado=None, bimestre=None):
    cand = buscar_alumno(estudiante, archivos, grado, bimestre)
    if not cand:
        return {"error": f'No encontré a "{estudiante}".'}
    if len(cand) > 1:
        return {"error": "ambiguo", "candidatos": cand}
    g, b, n, nom = cand[0]
    area = _match_area(curso)
    if not area:
        return {"error": f'No reconozco el curso "{curso}". Cursos disponibles: {", ".join(AREAS_HOJA)}'}
    hoja = AREAS_HOJA[area]
    nombre_reg = indexar_registros(archivos)["registros"][g][b]
    reg = _wb(nombre_reg, archivos)
    docente = reg[hoja]["D7"].value if hoja in reg.sheetnames else None
    cons = reg["CONSOLIDADO"]
    fila = n + 4
    mapa = _construir_mapa_areas()
    comp, notas_num = [], []
    for texto, cu1, cu2, cbim in mapa.get(area, []):
        u1 = cons.cell(fila, cu1).value if cu1 else None
        u2 = cons.cell(fila, cu2).value if cu2 else None
        bim = cons.cell(fila, cbim).value if cbim else None
        comp.append({"competencia": texto, "U1": u1 or "-", "U2": u2 or "-", "BIM": bim or "-"})
        if bim in _ESCALA:
            notas_num.append(_ESCALA[bim])
    promedio = round(sum(notas_num) / len(notas_num)) if notas_num else None
    return {
        "estudiante": nom, "grado": g, "bimestre": b, "area": area,
        "docente": (docente or "").strip() or "no registrado",
        "competencias": comp,
        "promedio_area": _ESCALA_INV.get(promedio, "-"),
    }


def resumen_notas_general(estudiante, archivos, grado=None, bimestre=None):
    cand = buscar_alumno(estudiante, archivos, grado, bimestre)
    if not cand:
        return {"error": f'No encontré a "{estudiante}".'}
    if len(cand) > 1:
        return {"error": "ambiguo", "candidatos": cand}
    g, b, n, nom = cand[0]
    nombre_reg = indexar_registros(archivos)["registros"][g][b]
    cons = _wb(nombre_reg, archivos)["CONSOLIDADO"]
    fila = n + 4
    mapa = _construir_mapa_areas()
    resumen, todas = [], []
    for area, items in mapa.items():
        notas_area = [_ESCALA[cons.cell(fila, cbim).value] for _, _, _, cbim in items
                      if cbim and cons.cell(fila, cbim).value in _ESCALA]
        if notas_area:
            prom = round(sum(notas_area) / len(notas_area))
            resumen.append({"area": area, "promedio": _ESCALA_INV.get(prom, "-")})
            todas.extend(notas_area)
    prom_general = round(sum(todas) / len(todas)) if todas else None
    return {
        "estudiante": nom, "grado": g, "bimestre": b, "por_area": resumen,
        "promedio_general": _ESCALA_INV.get(prom_general, "-"),
    }


def consultar_asistencia(estudiante, archivos, grado=None, bimestre=None):
    cand = buscar_alumno(estudiante, archivos, grado, bimestre)
    if not cand:
        return {"error": f'No encontré a "{estudiante}".'}
    if len(cand) > 1:
        return {"error": "ambiguo", "candidatos": cand}
    g, b, n, nom = cand[0]
    nombre_reg = indexar_registros(archivos)["registros"][g][b]
    reg = _wb(nombre_reg, archivos)
    if "ASISTEN" not in reg.sheetnames:
        return {"error": "Este registro no tiene hoja de asistencia."}
    ws = reg["ASISTEN"]
    fila = None
    for r in range(11, 50):
        if ws.cell(r, 1).value == n:
            fila = r
            break
    if fila is None:
        return {"error": f"No encontré la fila de asistencia del alumno N° {n}."}

    faltas, tardanzas, falt_just = [], [], []
    for c in range(12, 62):
        fecha = ws.cell(7, c).value
        if fecha is None:
            continue
        code = ws.cell(fila, c).value
        fstr = fecha.strftime("%d/%m/%Y")
        if code == "T":
            tardanzas.append(fstr)
        elif code == "F":
            faltas.append(fstr)
        elif code == "FJ":
            falt_just.append(fstr)

    pct = ws.cell(fila, 71).value
    return {
        "estudiante": nom, "grado": g, "bimestre": b,
        "total_dias_asistidos": ws.cell(fila, 67).value,
        "porcentaje_asistencia": f"{round(pct * 100, 1)}%" if isinstance(pct, (int, float)) else pct,
        "total_tardanzas": ws.cell(fila, 68).value, "fechas_tardanzas": tardanzas,
        "total_faltas_injustificadas": ws.cell(fila, 70).value, "fechas_faltas": faltas,
        "total_faltas_justificadas": ws.cell(fila, 69).value, "fechas_faltas_justificadas": falt_just,
    }
